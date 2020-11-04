# -*- coding: utf-8 -*-
import serial
import openpyxl
import os
import re
import tkinter
window = tkinter.Tk()
window.title("POS EXCEL 프로그램")
window.geometry("300x200+100+100")
window.resizable(False, False)
ser = serial.Serial("COM9", 9600)

def function() :
    menu = 0
    delivery = 0
    wb = openpyxl.load_workbook('C:\\Users\\User\\Documents\\score.xlsx')
    ws = wb.active
    address = 0
    time = 0
    row = 1
    flag = 0
    data = {}
    flag2 = 0
    request = 0
    company = 0
    while ser.readable() :
        x = ser.readline()
        #print(x)
        serialData = x.decode('utf-8', 'ignore')
        #print(data)
    #######################################################
        if "주문시간" in serialData :
            time = serialData
            print (time)
            row += 1

        if '메뉴명' in serialData :
            menu = 1
        if '배달비' in serialData :
            delivery = 1

        if "주소" in serialData :
            address = serialData
            print(address)
            flag = 1

        if "주문요청사항" in serialData :
            request = serialData
            print(request)

        if "행컵" in serialData:
            company = serialData
            print(company)

        if "워너비박스" in serialData:
            company = serialData
            print(company)

        if "믿고" in serialData:
            company = serialData
            print(company)

        if "이태리돈까스" in serialData:
            company = serialData
            print(company)

        if "오떡후" in serialData:
            company = serialData
            print(company)

    #여러 줄이 메뉴를 list에 저장시키고 string으로 변환시켜 필요없는 친구들을 제거하는 과정~~

        if menu == 1 :
            f = open("D:\output.txt", 'a', errors = 'ignore')
            f.write(serialData)

        if delivery == 1 :
            f.close()
            file = open('D:\output.txt', 'r')
            x = file.readlines()
            print(x)
            #d = ''.join(x)
            #extraction(x)  #나도 함수를 쓰고 싶다.. 코드 정리하고 싶다

            data = ' '.join(x).split()
            #d = list(x)
            print(type(data))
            print(data)
            data.remove("메뉴명")
            data.remove("수량")
            data.remove("금액")
            data.remove("-------------------------------")
            data.remove("배달비")
            print("쓸 데 없는 거 제거 : ", data)

            data = ' '.join(data) #문자열로 만들어주는 친구
            p = re.compile("[^0-9]")
            data = "".join(p.findall(data))
            print("숫자 제거 : ", data)

            data = list(data) #문자열을 한 글자씩 나누어서 list로 만들어주는 친구
            for dd in data :
                if "원" in data :
                    data.remove("원")
            print("원 제거 : ", data)
            data = ' '.join(data)  # 문자열로 만들어주는 친구
            print(data)

            #for i in d:
            #    d = ' '.join(i)  # 문자열로 만들어주는 친구
            #print(d)

            file.close()
            menu = 0
            delivery = 0
            os.remove("D:\output.txt")
            flag2 = 1


        if flag == 1 : #주소를 바로 출력시키기 위한
            for i in range(5) :
                if i == 0 :
                    w = address
                    ws.cell(row, 1, w)
                if i == 1 :
                    w = time
                    ws.cell(row, 2, w)

                if flag2 == 1 : #data의 string 오류를 방지하기 위해서
                    if i == 2 :
                        w = data
                        print(w)
                    #for i in d :
                       #print(i)
                        ws.cell(row, 3, w)
                        flag2 = 0
                if i == 3 :
                    w = request
                    ws.cell(row, 4, w)

                if i == 4 :
                    w = company
                    ws.cell(row, 5, w)

                if i == 4 :
                    i = 0
                flag = 0

        wb.save("C:\\Users\\User\\Documents\\score.xlsx")
        wb.close()

label = tkinter.Label(window, text = "")
label.pack()

label = tkinter.Label(window, text = "")
label.pack()

label = tkinter.Label(window, text = "")
label.pack()

button = tkinter.Button(window, overrelief="solid", width=30, command=function, repeatdelay=100, repeatinterval=100, text = "클릭하세용")
button.pack()

window.mainloop()